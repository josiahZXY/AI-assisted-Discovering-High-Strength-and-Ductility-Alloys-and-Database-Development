import openpyxl
from copy import deepcopy
def write_excel_xlsx(path, sheet_name, value):
	index = len(value)
	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = sheet_name
	for i in range(0, index):
		for j in range(0, len(value[i])):
			sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
	workbook.save(path)
	print("xlsx格式表格写入数据成功！")


def read_excel_xlsx(path, sheet_name):
	workbook = openpyxl.load_workbook(path)
	# sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
	sheet = workbook[sheet_name]
	for row in sheet.rows:
		for cell in row:
			print(cell.value, "\t", end="")


def save_data(excel_name, data_list):
	book_name_xlsx = excel_name + '.xlsx'
	
	sheet_name_xlsx = '当当网数据'
	excel_title = ["书名", "作者", "价格", "出版时间"]
	write_data = deepcopy(data_list)
	write_data.insert(0, excel_title)
	write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, write_data)
	read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)



if __name__=='__main__':
	f = open('data.txt', 'r')
	data_dict = f.readlines()
	data_list = [[eval(i)['书名'], eval(i)['作者'], eval(i)['价格'],  eval(i)['出版时间'] ] for i in data_dict]
	print(data_list)
	excel_name = input('请输入excel文件名称:')
	save_data(excel_name, data_list)